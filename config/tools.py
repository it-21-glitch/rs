import os
import uuid
import sqlite3
import shutil
import pandas as pd
import openpyxl as xl
from datetime import datetime
from openpyxl.styles import NamedStyle
from openpyxl.styles import Border, Side
from flask_wtf import FlaskForm
from wtforms.validators import DataRequired, Length
from wtforms import StringField, SubmitField, PasswordField, FileField


class DefaultConfig:  # 生产模式 所需要的全部参数
    SECRET_KEY = 'asdadwqedsadsadaqeasczxcxzcasfadasda'
    WTF_CSRF_SECRET_KEY = 'a random string'  # 使用不同的密钥


class DevelopmentConfig(DefaultConfig):  # 调试模式下的参数
    '''继承了生成模式的类'''
    DEBUG = True


class VerifyForm(FlaskForm):
    name = StringField('账户', validators=[DataRequired()])
    pwd = PasswordField('密码', validators=[DataRequired(), Length(min=8, max=128)])
    submit = SubmitField('登录')


class VerifyUpdateForm(FlaskForm):
    file = FileField('上传文件')
    submit = SubmitField('上传')


# 规则解析函数
def regular_function():
    parsed_data = []
    base_path = os.path.dirname(os.path.abspath(os.path.dirname(__file__)))
    path = os.path.join(base_path, 'static', 'rule', 'rule.xlsx')
    if not os.path.exists(path):
        return parsed_data
    df = pd.read_excel(path, header=3)  # 假设第三行是标题行
    # 对每一列进行空值填充，用上一行的值填充
    # df.fillna(method='ffill', inplace=True)
    df.ffill(inplace=True)
    # 遍历DataFrame的每一行
    for index, row in df.iterrows():
        list_row = []
        for value in row:
            if isinstance(value, float):
                value = round(value)  # 进行4舍5入
            list_row.append(value)
        # 将字典添加到列表中
        parsed_data.append(list_row)
    return parsed_data


# 存储规则记录
def rule_data_base(rule_data_list, cursor, db_conn):
    status = True
    del_sql = """
        DELETE FROM main.material;
        DELETE FROM main.process;
        DELETE FROM main.equipment;
    """
    try:
        cursor.executescript(del_sql)  # 使用executescript()执行多条SQL语句
        db_conn.commit()
    except sqlite3.Error as e:
        # 如果发生错误，回滚事务
        db_conn.rollback()
        status = False
        return status
    # 存储规则
    for i in rule_data_list:
        # 添加材质/查询材质
        material_name = i[0]
        material_select_sql = f"""
                SELECT id FROM main.material WHERE material_name='{material_name}'
            """
        cursor.execute(material_select_sql)
        material_info = cursor.fetchone()
        if not material_info:
            material_insert_sql = f"""
                    INSERT INTO main.material(material_name) VALUES('{material_name}')
                """
            try:
                cursor.execute(material_insert_sql)
                db_conn.commit()
            except sqlite3.Error as e:
                # 如果发生错误，回滚事务
                db_conn.rollback()
                status = False
                break
            material_id = cursor.lastrowid
        else:
            material_id = material_info[0]
        process_name = i[1]
        # 添加工序/查询工序
        process_select_sql = f"""
                SELECT id FROM main.process WHERE process_name='{process_name}'
            """
        cursor.execute(process_select_sql)
        process_info = cursor.fetchone()
        if not process_info:
            process_insert_sql = f"""
                    INSERT INTO main.process(process_name,material_id) VALUES('{process_name}','{material_id}')
                """
            try:
                cursor.execute(process_insert_sql)
                db_conn.commit()
            except sqlite3.Error as e:
                db_conn.rollback()
                status = False
                break
            process_id = cursor.lastrowid
        else:
            process_id = process_info[0]
        # 查看设备和次数/添加设备和次数
        equipment_name = i[2]
        equipment_number = i[3]
        max_people_number = i[5]
        min_people_number = i[4]

        day_classes_frequency = i[6]
        classes_man_hour = i[7]
        day_classes_man_hour = i[8]
        classes_capacity_big = i[9]
        classes_capacity_middle = i[10]
        classes_capacity_small = i[11]

        day_capacity_big = i[12]
        day_capacity_middle = i[13]
        day_capacity_small = i[14]

        equipment_select_sql = f"""
                SELECT id FROM main.equipment WHERE process_id='{process_id}' AND equipment_name='{equipment_name}'
            """
        cursor.execute(equipment_select_sql)
        equipment_info = cursor.fetchone()
        if not equipment_info:
            equipment_insert_sql = f"""
                INSERT INTO 
                    main.equipment(
                        process_id,
                        equipment_name,
                        equipment_number,
                        max_people_number,
                        min_people_number,
                        
                        classes_capacity_big,
                        classes_capacity_middle,
                        classes_capacity_small,
                        
                        day_classes_frequency,
                        classes_man_hour,
                        day_classes_man_hour,
                        
                        day_capacity_big,
                        day_capacity_middle,
                        day_capacity_small
                        ) 
                VALUES
                    (
                        '{process_id}',
                        '{equipment_name}',
                        '{equipment_number}',
                        '{max_people_number}',
                        '{min_people_number}',
                        '{classes_capacity_big}',
                        '{classes_capacity_middle}',
                        '{classes_capacity_small}',
                        '{day_classes_frequency}',
                        '{classes_man_hour}',
                        '{day_classes_man_hour}',
                        '{day_capacity_big}',
                        '{day_capacity_middle}',
                        '{day_capacity_small}'
                    )
                """
            try:
                cursor.execute(equipment_insert_sql)
                db_conn.commit()
            except sqlite3.Error as e:
                db_conn.rollback()
                status = False
                break
    cursor.close()
    db_conn.close()
    return status


# 模板过滤器方法
def audit_status(data):  # 注册过滤器的函数
    '''
    :param data:  接收模板传入的参数
    :return:
    '''
    if not data:
        return "未审核"
    return "已审核"


# 表格生成
def generation_xlsx(pk, get_db):
    file_dir = uuid.uuid4()
    base_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static')
    template_file = os.path.join(base_path, 'template_xlsx', 'Template.xlsx')  # 薪酬
    template_file_1 = os.path.join(base_path, 'template_xlsx', 'Template_1.xlsx')  # 产能记录
    db_conn = get_db()
    cursor = db_conn.cursor()
    wb = xl.load_workbook(template_file)
    wb_1 = xl.load_workbook(template_file_1)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]  # 第一个表薪酬
    ws2_by_index = wb_1.active  # 第二个 产能记录

    styleFont = xl.styles.Font(name='Arial', size=10)  # 字体
    styleAlignment = xl.styles.Alignment(horizontal='center', vertical='center')  # 居中
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # 考勤记录和薪酬记录
    get_sql = f"""
        SELECT 
            fau.start_time,
            fau.end_time,
            fu.user_name,
            fau.user_pay,
            fau.classs_number
        FROM 
            main.factory_attendance_user  AS fau
        INNER JOIN 
            main.factory_user AS fu
        ON 
            fau.user_id= fu.id
        WHERE fau.record_sheet_id = '{pk}'
    """
    cursor.execute(get_sql)
    column_names = [description[0] for description in cursor.description]
    data_list_all = [dict(zip(column_names, row)) for row in cursor.fetchall()]
    # 数据去重
    uniq_list = []
    new_user_work_attendance_list = []
    for i in data_list_all:
        user_name = i.get("user_name")
        classs_number = i.get("classs_number")
        uniq_str = f'{user_name}-{classs_number}'
        if uniq_str in uniq_list:
            continue
        uniq_list.append(uniq_str)
        new_user_work_attendance_list.append(i)
    # 考勤
    ws['A1'] = ws['A1'].value + ' ' + new_user_work_attendance_list[0].get("start_time").split(' ')[0]  # 对名字设置日期
    new_user_work_attendance_dict = {}
    for i in new_user_work_attendance_list:
        if not new_user_work_attendance_dict.get(i.get("user_name")):
            new_user_work_attendance_dict[i.get("user_name")] = [i.get("start_time"), i.get("end_time")]
        else:
            new_user_work_attendance_dict[i.get("user_name")].append(i.get("start_time"))
            new_user_work_attendance_dict[i.get("user_name")].append(i.get("end_time"))
    # 获取最长的列（默认确定是最完整的）
    # 获取最长的列（默认确定是最完整的）
    max_length = max(len(times) for times in new_user_work_attendance_dict.values())
    max_length_data = []
    for name, times in new_user_work_attendance_dict.items():
        while len(times) < max_length:
            times.append('')
        if len(times) == max_length and '' not in times:
            max_length_data = times
        times.sort()
    for k, v in new_user_work_attendance_dict.items():
        if "" not in v:
            continue
        num = 0
        for index, i in enumerate(v):
            if not i:
                continue
            while len(max_length_data) > num:
                if max_length_data[num] == i:
                    v[num] = i
                    v[index] = ''
                    break
                num += 1
    # 样式边框
    title_row = 3
    title_column = 3
    for i in range(0, max_length):  # 设置班次标题
        line_letter = xl.utils.get_column_letter(title_column)  # 根据数字转换为列的坐标 起始
        line_letter_next = xl.utils.get_column_letter(title_column + 1)  # 根据数字转换为列的坐标 起始坐标+1
        if i % 2:
            ws[f'{line_letter}{title_row}'] = f'班次-签退'
        else:
            ws[f'{line_letter}{title_row}'] = f'班次-签到'
        ws[f'{line_letter}{title_row}'].alignment = styleAlignment
        ws.merge_cells(f'{line_letter}{title_row}:{line_letter_next}{title_row + 1}')
        title_column += 2
    ws.merge_cells(f'A1:{xl.utils.get_column_letter(title_column - 1)}2')  # 标题合并

    work_attendance_row_num = 5
    for key, val in new_user_work_attendance_dict.items():
        work_attendance_column_num = 3
        user_name = key
        ws[f'A{work_attendance_row_num}'] = user_name
        ws[f'A{work_attendance_row_num}'].font = styleFont
        ws[f'A{work_attendance_row_num}'].alignment = styleAlignment
        ws.merge_cells(f'A{work_attendance_row_num}:B{work_attendance_row_num + 1}')
        for i in val:
            line_letter = xl.utils.get_column_letter(work_attendance_column_num)  # 根据数字转换为列的坐标 起始
            line_letter_next = xl.utils.get_column_letter(work_attendance_column_num + 1)  # 根据数字转换为列的坐标 起始坐标+1
            ws[f'{line_letter}{work_attendance_row_num}'] = i
            # 设置字体样式
            ws[f'{line_letter}{work_attendance_row_num}'].font = styleFont
            ws[f'{line_letter}{work_attendance_row_num}'].alignment = styleAlignment
            # 合并居中
            ws.merge_cells(f'{line_letter}{work_attendance_row_num}:{line_letter_next}{work_attendance_row_num + 1}')
            work_attendance_column_num += 2
        work_attendance_row_num += 2  # 每两格子区分一下
    # 设置合并单元格的边框
    for row in ws.iter_rows(min_row=3, max_row=work_attendance_row_num - 1, min_col=1, max_col=title_column - 1):
        for cell in row:
            cell.border = thin_border

    # 产能
    row_num = 5
    get_record_sheet_sql = f"""
        SELECT
            rs.po,
            rs.item,
            rs.entry_time,
            rs.specification_name,
            rs.material_name,
            rs.process_name,
            rs.equipment_name,
            rs.people_number,
            rs.output_number
        FROM
            main.record_sheet  AS rs
        WHERE rs.id = '{pk}'
    """
    cursor.execute(get_record_sheet_sql)
    column_names = [description[0] for description in cursor.description]
    data_dict = dict(zip(column_names, cursor.fetchone()))
    po = data_dict.get("po")
    item = data_dict.get("item")
    entry_time = data_dict.get("entry_time")
    specification_name = data_dict.get("specification_name")
    material_name = data_dict.get("material_name")
    process_name = data_dict.get("process_name")
    equipment_name = data_dict.get("equipment_name")
    people_number = data_dict.get("people_number")
    output_number = data_dict.get("output_number")
    ws2_by_index[f'A{row_num}'] = entry_time
    ws2_by_index[f'C{row_num}'] = po
    ws2_by_index[f'E{row_num}'] = item
    ws2_by_index[f'G{row_num}'] = material_name
    ws2_by_index[f'I{row_num}'] = process_name
    ws2_by_index[f'K{row_num}'] = equipment_name
    ws2_by_index[f'M{row_num}'] = people_number
    ws2_by_index[f'O{row_num}'] = specification_name
    ws2_by_index[f'Q{row_num}'] = output_number
    ws2_by_index.merge_cells(f'A{row_num}:B{row_num + 1}')
    ws2_by_index.merge_cells(f'C{row_num}:D{row_num + 1}')
    ws2_by_index.merge_cells(f'E{row_num}:F{row_num + 1}')
    ws2_by_index.merge_cells(f'G{row_num}:H{row_num + 1}')
    ws2_by_index.merge_cells(f'I{row_num}:J{row_num + 1}')
    ws2_by_index.merge_cells(f'K{row_num}:L{row_num + 1}')
    ws2_by_index.merge_cells(f'M{row_num}:N{row_num + 1}')
    ws2_by_index.merge_cells(f'O{row_num}:P{row_num + 1}')
    ws2_by_index.merge_cells(f'Q{row_num}:R{row_num + 1}')
    ws2_by_index[f'A{row_num}'].font = styleFont
    ws2_by_index[f'A{row_num}'].alignment = styleAlignment
    ws2_by_index[f'C{row_num}'].font = styleFont
    ws2_by_index[f'C{row_num}'].alignment = styleAlignment
    ws2_by_index[f'E{row_num}'].font = styleFont
    ws2_by_index[f'E{row_num}'].alignment = styleAlignment
    ws2_by_index[f'G{row_num}'].font = styleFont
    ws2_by_index[f'G{row_num}'].alignment = styleAlignment
    ws2_by_index[f'I{row_num}'].font = styleFont
    ws2_by_index[f'I{row_num}'].alignment = styleAlignment
    ws2_by_index[f'K{row_num}'].font = styleFont
    ws2_by_index[f'K{row_num}'].alignment = styleAlignment
    ws2_by_index[f'M{row_num}'].font = styleFont
    ws2_by_index[f'M{row_num}'].alignment = styleAlignment
    ws2_by_index[f'O{row_num}'].font = styleFont
    ws2_by_index[f'O{row_num}'].alignment = styleAlignment
    ws2_by_index[f'Q{row_num}'].font = styleFont
    ws2_by_index[f'Q{row_num}'].alignment = styleAlignment
    for row in ws2_by_index.iter_rows(min_row=5, max_row=row_num + 1, min_col=1, max_col=19 - 1):
        for cell in row:
            cell.border = thin_border
    dir_path = os.path.join(base_path, 'generation_xlsx', f'{file_dir}')
    if not os.path.exists(dir_path):
        os.mkdir(dir_path)

    file_path = os.path.join(dir_path, '考勤记录.xlsx')
    file_path_1 = os.path.join(dir_path, '生产记录.xlsx')
    wb.save(file_path)
    wb_1.save(file_path_1)

    cursor.close()
    db_conn.close()

    shutil.make_archive(base_name=dir_path, format="zip", root_dir=dir_path)
    if os.path.exists(f"{dir_path}.zip"):
        shutil.rmtree(dir_path)
    return f"{dir_path}.zip"


def attendance_xlsx(data):
    file_name = uuid.uuid4()
    base_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static')
    template_file = os.path.join(base_path, 'template_xlsx', 'Template.xlsx')  # 薪酬
    user = data[0].get("user_name")
    data_list_all = data
    uniq_list = []
    new_user_work_attendance_list = []
    for i in data_list_all:
        user_name = i.get("user_name")
        classs_number = i.get("classs_number")
        start_time_day = i.get("start_time").split(" ")[0].split('-')[-1]
        start_time_month = i.get("start_time").split(" ")[0].split('-')[-2]
        uniq_str = f'{user_name}-{classs_number}-{start_time_month}-{start_time_day}'
        if uniq_str in uniq_list:
            continue
        uniq_list.append(uniq_str)
        new_user_work_attendance_list.append(i)
    # 考勤
    new_user_work_attendance_dict = {}
    for i in new_user_work_attendance_list:
        start_time_day = i.get("start_time").split(" ")[0].split('-')[-1]
        start_time_month = i.get("start_time").split(" ")[0].split('-')[-2]
        name_day = f'{i.get("user_name")}-{start_time_month}-{start_time_day}'
        if not new_user_work_attendance_dict.get(name_day):
            new_user_work_attendance_dict[name_day] = [i.get("start_time"), i.get("end_time")]
        else:
            new_user_work_attendance_dict[name_day].append(i.get("start_time"))
            new_user_work_attendance_dict[name_day].append(i.get("end_time"))
    # 获取最长的列（默认确定是最完整的）
    # 获取最长的列（默认确定是最完整的）
    max_length = max(len(times) for times in new_user_work_attendance_dict.values())
    max_length_data = []
    for name, times in new_user_work_attendance_dict.items():
        while len(times) < max_length:
            times.append('')
        if len(times) == max_length and '' not in times:
            max_length_data = times
        times.sort()
    for k, v in new_user_work_attendance_dict.items():
        if "" not in v:
            continue
        num = 0
        for index, i in enumerate(v):
            if not i:
                continue
            while len(max_length_data) > num:
                if max_length_data[num] == i:
                    v[num] = i
                    v[index] = ''
                    break
                num += 1

    wb = xl.load_workbook(template_file)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]  # 第一个表薪酬
    styleFont = xl.styles.Font(name='Arial', size=10)  # 字体
    styleAlignment = xl.styles.Alignment(horizontal='center', vertical='center')  # 居中
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    title_row = 3
    title_column = 3
    for i in range(0, max_length):  # 设置班次标题
        line_letter = xl.utils.get_column_letter(title_column)  # 根据数字转换为列的坐标 起始
        line_letter_next = xl.utils.get_column_letter(title_column + 1)  # 根据数字转换为列的坐标 起始坐标+1
        if i % 2:
            ws[f'{line_letter}{title_row}'] = f'班次-签退'
        else:
            ws[f'{line_letter}{title_row}'] = f'班次-签到'
        ws[f'{line_letter}{title_row}'].alignment = styleAlignment
        ws.merge_cells(f'{line_letter}{title_row}:{line_letter_next}{title_row + 1}')
        title_column += 2
    ws.merge_cells(f'A1:{xl.utils.get_column_letter(title_column - 1)}2')  # 标题合并

    work_attendance_row_num = 5
    for key, val in new_user_work_attendance_dict.items():
        work_attendance_column_num = 3
        user_name = key
        ws[f'A{work_attendance_row_num}'] = user_name
        ws[f'A{work_attendance_row_num}'].font = styleFont
        ws[f'A{work_attendance_row_num}'].alignment = styleAlignment
        ws.merge_cells(f'A{work_attendance_row_num}:B{work_attendance_row_num + 1}')
        for i in val:
            line_letter = xl.utils.get_column_letter(work_attendance_column_num)  # 根据数字转换为列的坐标 起始
            line_letter_next = xl.utils.get_column_letter(work_attendance_column_num + 1)  # 根据数字转换为列的坐标 起始坐标+1
            ws[f'{line_letter}{work_attendance_row_num}'] = i
            # 设置字体样式
            ws[f'{line_letter}{work_attendance_row_num}'].font = styleFont
            ws[f'{line_letter}{work_attendance_row_num}'].alignment = styleAlignment
            # 合并居中
            ws.merge_cells(f'{line_letter}{work_attendance_row_num}:{line_letter_next}{work_attendance_row_num + 1}')
            work_attendance_column_num += 2
        work_attendance_row_num += 2  # 每两格子区分一下
    # 设置合并单元格的边框
    for row in ws.iter_rows(min_row=3, max_row=work_attendance_row_num - 1, min_col=1, max_col=title_column - 1):
        for cell in row:
            cell.border = thin_border
    file_name = f'{file_name}.xlsx'
    file_path = os.path.join(base_path, 'generation_xlsx', file_name)
    wb.save(file_path)

    return file_name


def produce_xlsx(data):
    file_name = uuid.uuid4()
    base_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static')
    template_file_1 = os.path.join(base_path, 'template_xlsx', 'Template_1.xlsx')  # 产能记录
    wb_1 = xl.load_workbook(template_file_1)
    ws2_by_index = wb_1.active  # 第二个 产能记录

    styleFont = xl.styles.Font(name='Arial', size=10)  # 字体
    styleAlignment = xl.styles.Alignment(horizontal='center', vertical='center')  # 居中
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    row_num = 5
    for data_dict in data:
        po = data_dict.get("po")
        item = data_dict.get("item")
        entry_time = data_dict.get("entry_time")
        specification_name = data_dict.get("specification_name")
        material_name = data_dict.get("material_name")
        process_name = data_dict.get("process_name")
        equipment_name = data_dict.get("equipment_name")
        people_number = data_dict.get("people_number")
        output_number = data_dict.get("output_number")
        ws2_by_index[f'A{row_num}'] = entry_time
        ws2_by_index[f'C{row_num}'] = po
        ws2_by_index[f'E{row_num}'] = item
        ws2_by_index[f'G{row_num}'] = material_name
        ws2_by_index[f'I{row_num}'] = process_name
        ws2_by_index[f'K{row_num}'] = equipment_name
        ws2_by_index[f'M{row_num}'] = people_number
        ws2_by_index[f'O{row_num}'] = specification_name
        ws2_by_index[f'Q{row_num}'] = output_number
        ws2_by_index.merge_cells(f'A{row_num}:B{row_num + 1}')
        ws2_by_index.merge_cells(f'C{row_num}:D{row_num + 1}')
        ws2_by_index.merge_cells(f'E{row_num}:F{row_num + 1}')
        ws2_by_index.merge_cells(f'G{row_num}:H{row_num + 1}')
        ws2_by_index.merge_cells(f'I{row_num}:J{row_num + 1}')
        ws2_by_index.merge_cells(f'K{row_num}:L{row_num + 1}')
        ws2_by_index.merge_cells(f'M{row_num}:N{row_num + 1}')
        ws2_by_index.merge_cells(f'O{row_num}:P{row_num + 1}')
        ws2_by_index.merge_cells(f'Q{row_num}:R{row_num + 1}')
        ws2_by_index[f'A{row_num}'].font = styleFont
        ws2_by_index[f'A{row_num}'].alignment = styleAlignment
        ws2_by_index[f'C{row_num}'].font = styleFont
        ws2_by_index[f'C{row_num}'].alignment = styleAlignment
        ws2_by_index[f'E{row_num}'].font = styleFont
        ws2_by_index[f'E{row_num}'].alignment = styleAlignment
        ws2_by_index[f'G{row_num}'].font = styleFont
        ws2_by_index[f'G{row_num}'].alignment = styleAlignment
        ws2_by_index[f'I{row_num}'].font = styleFont
        ws2_by_index[f'I{row_num}'].alignment = styleAlignment
        ws2_by_index[f'K{row_num}'].font = styleFont
        ws2_by_index[f'K{row_num}'].alignment = styleAlignment
        ws2_by_index[f'M{row_num}'].font = styleFont
        ws2_by_index[f'M{row_num}'].alignment = styleAlignment
        ws2_by_index[f'O{row_num}'].font = styleFont
        ws2_by_index[f'O{row_num}'].alignment = styleAlignment
        ws2_by_index[f'Q{row_num}'].font = styleFont
        ws2_by_index[f'Q{row_num}'].alignment = styleAlignment
        for row in ws2_by_index.iter_rows(min_row=5, max_row=row_num + 1, min_col=1, max_col=19 - 1):
            for cell in row:
                cell.border = thin_border
        row_num += 2

    file_name = f'{file_name}.xlsx'
    file_path = os.path.join(base_path, 'generation_xlsx', file_name)
    wb_1.save(file_path)

    return file_name
