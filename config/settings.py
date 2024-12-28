import os
import uuid
import sqlite3
import shutil
import pandas as pd
import openpyxl as xl
from datetime import datetime
from openpyxl.styles import NamedStyle
from openpyxl.styles import Border, Side
from flask_wtf import FlaskForm
from wtforms.validators import DataRequired, Length
from wtforms import StringField, SubmitField, PasswordField, FileField


class DefaultConfig:  # 生产模式 所需要的全部参数
    SECRET_KEY = 'asdadwqedsadsadaqeasczxcxzcasfadasda'
    WTF_CSRF_SECRET_KEY = 'a random string'  # 使用不同的密钥


class DevelopmentConfig(DefaultConfig):  # 调试模式下的参数
    '''继承了生成模式的类'''
    DEBUG = True


class VerifyForm(FlaskForm):
    name = StringField('账户', validators=[DataRequired()])
    pwd = PasswordField('密码', validators=[DataRequired(), Length(min=8, max=128)])
    submit = SubmitField('登录')


class VerifyUpdateForm(FlaskForm):
    file = FileField('上传文件')
    submit = SubmitField('上传')


# 规则解析函数
def regular_function():
    parsed_data = []
    base_path = os.path.dirname(os.path.abspath(os.path.dirname(__file__)))
    path = os.path.join(base_path, 'static', 'rule', 'rule.xlsx')
    if not os.path.exists(path):
        return parsed_data
    df = pd.read_excel(path, header=3)  # 假设第三行是标题行
    # 对每一列进行空值填充，用上一行的值填充
    # df.fillna(method='ffill', inplace=True)
    df.ffill(inplace=True)
    # 遍历DataFrame的每一行
    for index, row in df.iterrows():
        list_row = []
        for value in row:
            if isinstance(value, float):
                value = round(value)  # 进行4舍5入
            list_row.append(value)
        # 将字典添加到列表中
        parsed_data.append(list_row)
    return parsed_data


# 存储规则记录
def rule_data_base(rule_data_list, cursor, db_conn):
    status = True
    del_sql = """
        DELETE FROM main.material;
        DELETE FROM main.process;
        DELETE FROM main.equipment;
    """
    try:
        cursor.executescript(del_sql)  # 使用executescript()执行多条SQL语句
        db_conn.commit()
    except sqlite3.Error as e:
        # 如果发生错误，回滚事务
        db_conn.rollback()
        status = False
        return status
    # 存储规则
    for i in rule_data_list:
        # 添加材质/查询材质
        material_name = i[0]
        material_select_sql = f"""
                SELECT id FROM main.material WHERE material_name='{material_name}'
            """
        cursor.execute(material_select_sql)
        material_info = cursor.fetchone()
        if not material_info:
            material_insert_sql = f"""
                    INSERT INTO main.material(material_name) VALUES('{material_name}')
                """
            try:
                cursor.execute(material_insert_sql)
                db_conn.commit()
            except sqlite3.Error as e:
                # 如果发生错误，回滚事务
                db_conn.rollback()
                status = False
                break
            material_id = cursor.lastrowid
        else:
            material_id = material_info[0]
        process_name = i[1]
        # 添加工序/查询工序
        process_select_sql = f"""
                SELECT id FROM main.process WHERE process_name='{process_name}'
            """
        cursor.execute(process_select_sql)
        process_info = cursor.fetchone()
        if not process_info:
            process_insert_sql = f"""
                    INSERT INTO main.process(process_name,material_id) VALUES('{process_name}','{material_id}')
                """
            try:
                cursor.execute(process_insert_sql)
                db_conn.commit()
            except sqlite3.Error as e:
                db_conn.rollback()
                status = False
                break
            process_id = cursor.lastrowid
        else:
            process_id = process_info[0]
        # 查看设备和次数/添加设备和次数
        equipment_name = i[2]
        equipment_number = i[3]
        max_people_number = i[5]
        min_people_number = i[4]

        day_classes_frequency = i[6]
        classes_man_hour = i[7]
        day_classes_man_hour = i[8]
        classes_capacity_big = i[9]
        classes_capacity_middle = i[10]
        classes_capacity_small = i[11]

        day_capacity_big = i[12]
        day_capacity_middle = i[13]
        day_capacity_small = i[14]

        equipment_select_sql = f"""
                SELECT id FROM main.equipment WHERE process_id='{process_id}' AND equipment_name='{equipment_name}'
            """
        cursor.execute(equipment_select_sql)
        equipment_info = cursor.fetchone()
        if not equipment_info:
            equipment_insert_sql = f"""
                INSERT INTO 
                    main.equipment(
                        process_id,
                        equipment_name,
                        equipment_number,
                        max_people_number,
                        min_people_number,
                        
                        classes_capacity_big,
                        classes_capacity_middle,
                        classes_capacity_small,
                        
                        day_classes_frequency,
                        classes_man_hour,
                        day_classes_man_hour,
                        
                        day_capacity_big,
                        day_capacity_middle,
                        day_capacity_small
                        ) 
                VALUES
                    (
                        '{process_id}',
                        '{equipment_name}',
                        '{equipment_number}',
                        '{max_people_number}',
                        '{min_people_number}',
                        '{classes_capacity_big}',
                        '{classes_capacity_middle}',
                        '{classes_capacity_small}',
                        '{day_classes_frequency}',
                        '{classes_man_hour}',
                        '{day_classes_man_hour}',
                        '{day_capacity_big}',
                        '{day_capacity_middle}',
                        '{day_capacity_small}'
                    )
                """
            try:
                cursor.execute(equipment_insert_sql)
                db_conn.commit()
            except sqlite3.Error as e:
                db_conn.rollback()
                status = False
                break
    cursor.close()
    db_conn.close()
    return status


# 模板过滤器方法
def audit_status(data):  # 注册过滤器的函数
    '''
    :param data:  接收模板传入的参数
    :return:
    '''
    if not data:
        return "未审核"
    return "已审核"


# 表格生成
def generation_xlsx(pk, get_db):
    file_dir = uuid.uuid4()
    base_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static')
    template_file = os.path.join(base_path, 'template_xlsx', 'Template.xlsx')  # 薪酬
    template_file_1 = os.path.join(base_path, 'template_xlsx', 'Template_1.xlsx')  # 产能记录
    # template_file_2 = os.path.join(base_path, 'template_xlsx', 'Template_2.xlsx')  # 考勤
    db_conn = get_db()
    cursor = db_conn.cursor()

    # 考勤记录和薪酬记录
    get_sql = f"""
        SELECT 
            fau.start_time,
            fau.end_time,
            fu.user_name,
            fau.user_pay 
        FROM 
            main.factory_attendance_user  AS fau
        INNER JOIN 
            main.factory_user AS fu
        ON 
            fau.user_id= fu.id
        WHERE fau.record_sheet_id = '{pk}'
    """
    cursor.execute(get_sql)
    column_names = [description[0] for description in cursor.description]
    data_list_all = [dict(zip(column_names, row)) for row in cursor.fetchall()]
    wb = xl.load_workbook(template_file)
    wb_1 = xl.load_workbook(template_file_1)
    # wb_2 = xl.load_workbook(template_file_2)
    # ws = wb.active  # 第一个表薪酬
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]  # 第一个表薪酬
    ws2_by_index = wb_1.active  # 第二个 产能记录

    styleFont = xl.styles.Font(name='Arial', size=10)  # 字体
    styleAlignment = xl.styles.Alignment(horizontal='center', vertical='center')  # 居中
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    row_num = 5
    row_interval_num = 2
    # 考勤
    for index, item in enumerate(data_list_all):
        user_name = item.get("user_name")
        start_time = item.get("start_time")
        end_time = item.get("end_time")
        user_pay = item.get("user_pay")
        start_time = datetime.strptime(start_time, '%Y-%m-%d %H:%M')
        end_time = datetime.strptime(end_time, '%Y-%m-%d %H:%M')
        time_difference = end_time - start_time
        hours_difference = round(time_difference.total_seconds() / 3600, 1)
        if index != 0:
            row_num += row_interval_num
        ws[f'A{row_num}'] = user_name
        ws[f'C{row_num}'] = start_time
        ws[f'E{row_num}'] = end_time
        ws[f'G{row_num}'] = hours_difference
        ws[f'I{row_num}'] = user_pay
        ws.merge_cells(f'A{row_num}:B{row_num + 1}')
        ws.merge_cells(f'C{row_num}:D{row_num + 1}')
        ws.merge_cells(f'E{row_num}:F{row_num + 1}')
        ws.merge_cells(f'G{row_num}:H{row_num + 1}')
        ws.merge_cells(f'I{row_num}:I{row_num + 1}')
        # 设置合并单元格的样式
        ws[f'A{row_num}'].font = styleFont
        ws[f'A{row_num}'].alignment = styleAlignment
        ws[f'C{row_num}'].font = styleFont
        ws[f'C{row_num}'].alignment = styleAlignment
        ws[f'E{row_num}'].font = styleFont
        ws[f'E{row_num}'].alignment = styleAlignment
        ws[f'G{row_num}'].font = styleFont
        ws[f'G{row_num}'].alignment = styleAlignment
        ws[f'I{row_num}'].font = styleFont
        ws[f'I{row_num}'].alignment = styleAlignment
        # 设置合并单元格的边框
        for row in ws.iter_rows(min_row=5, max_row=row_num + 1, min_col=1, max_col=10 - 1):
            for cell in row:
                cell.border = thin_border

    # row_num = 5
    # row_interval_num = 2
    # # 薪酬
    # for index, item in enumerate(data_list_all):
    #     user_name = item.get("user_name")
    #     user_pay = item.get("user_pay")
    #     if index != 0:
    #         row_num += row_interval_num
    #     ws[f'A{row_num}'] = user_name
    #     ws[f'D{row_num}'] = user_pay
    #     ws.merge_cells(f'A{row_num}:C{row_num + 1}')
    #     ws.merge_cells(f'D{row_num}:F{row_num + 1}')
    #     # 设置合并单元格的样式
    #     ws[f'A{row_num}'].font = styleFont
    #     ws[f'A{row_num}'].alignment = styleAlignment
    #     ws[f'D{row_num}'].font = styleFont
    #     ws[f'D{row_num}'].alignment = styleAlignment
    #     # 设置合并单元格的边框
    #     for row in ws.iter_rows(min_row=5, max_row=row_num + 1, min_col=1, max_col=7 - 1):
    #         for cell in row:
    #             cell.border = thin_border

    # 产能
    row_num = 5
    get_record_sheet_sql = f"""
        SELECT 
            rs.po,
            rs.item,
            rs.specification_name,
            rs.material_name,
            rs.process_name,
            rs.equipment_name,
            rs.people_number,
            rs.output_number
        FROM 
            main.record_sheet  AS rs
        WHERE rs.id = '{pk}'
    """
    cursor.execute(get_record_sheet_sql)
    column_names = [description[0] for description in cursor.description]
    data_dict = dict(zip(column_names, cursor.fetchone()))
    po = data_dict.get("po")
    item = data_dict.get("item")
    specification_name = data_dict.get("specification_name")
    material_name = data_dict.get("material_name")
    process_name = data_dict.get("process_name")
    equipment_name = data_dict.get("equipment_name")
    people_number = data_dict.get("people_number")
    output_number = data_dict.get("output_number")
    ws2_by_index[f'A{row_num}'] = po
    ws2_by_index[f'C{row_num}'] = item
    ws2_by_index[f'E{row_num}'] = material_name
    ws2_by_index[f'G{row_num}'] = process_name
    ws2_by_index[f'I{row_num}'] = equipment_name
    ws2_by_index[f'k{row_num}'] = people_number
    ws2_by_index[f'M{row_num}'] = specification_name
    ws2_by_index[f'O{row_num}'] = output_number
    ws2_by_index.merge_cells(f'A{row_num}:B{row_num + 1}')
    ws2_by_index.merge_cells(f'C{row_num}:D{row_num + 1}')
    ws2_by_index.merge_cells(f'E{row_num}:F{row_num + 1}')
    ws2_by_index.merge_cells(f'G{row_num}:H{row_num + 1}')
    ws2_by_index.merge_cells(f'I{row_num}:J{row_num + 1}')
    ws2_by_index.merge_cells(f'K{row_num}:L{row_num + 1}')
    ws2_by_index.merge_cells(f'M{row_num}:N{row_num + 1}')
    ws2_by_index.merge_cells(f'O{row_num}:P{row_num + 1}')
    ws2_by_index[f'A{row_num}'].font = styleFont
    ws2_by_index[f'A{row_num}'].alignment = styleAlignment
    ws2_by_index[f'C{row_num}'].font = styleFont
    ws2_by_index[f'C{row_num}'].alignment = styleAlignment
    ws2_by_index[f'E{row_num}'].font = styleFont
    ws2_by_index[f'E{row_num}'].alignment = styleAlignment
    ws2_by_index[f'G{row_num}'].font = styleFont
    ws2_by_index[f'G{row_num}'].alignment = styleAlignment
    ws2_by_index[f'I{row_num}'].font = styleFont
    ws2_by_index[f'I{row_num}'].alignment = styleAlignment
    ws2_by_index[f'K{row_num}'].font = styleFont
    ws2_by_index[f'K{row_num}'].alignment = styleAlignment
    ws2_by_index[f'M{row_num}'].font = styleFont
    ws2_by_index[f'M{row_num}'].alignment = styleAlignment
    ws2_by_index[f'O{row_num}'].font = styleFont
    ws2_by_index[f'O{row_num}'].alignment = styleAlignment
    for row in ws2_by_index.iter_rows(min_row=5, max_row=row_num + 1, min_col=1, max_col=17 - 1):
        for cell in row:
            cell.border = thin_border
    dir_path = os.path.join(base_path, 'generation_xlsx', f'{file_dir}')
    if not os.path.exists(dir_path):
        os.mkdir(dir_path)

    file_path = os.path.join(dir_path, '考勤薪酬记录.xlsx')
    file_path_1 = os.path.join(dir_path, '生产记录.xlsx')
    # file_path_2 = os.path.join(dir_path, '考勤.xlsx')
    wb.save(file_path)
    wb_1.save(file_path_1)
    # wb_2.save(file_path_2)
    cursor.close()
    db_conn.close()

    shutil.make_archive(base_name=dir_path, format="zip", root_dir=dir_path)
    if os.path.exists(f"{dir_path}.zip"):
        shutil.rmtree(dir_path)
    return f"{dir_path}.zip"
